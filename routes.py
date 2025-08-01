import os
from flask import render_template, request, redirect, url_for, flash, jsonify, send_file, session
from werkzeug.utils import secure_filename
from flask_login import login_user, logout_user, login_required, current_user
from flask_wtf.csrf import generate_csrf
from app import app, db
from models import Company, Customer, Product, Bill, BillItem, BillSequence, User, Category
from forms import CompanyConfigForm, CustomerForm, ProductForm, BillForm, BillItemForm, LoginForm, UserForm, ChangePasswordForm, CreateUserForm, QuickAddProductForm
from utils import allowed_file, get_state_name
from gst_calculator import calculate_gst
from pdf_generator import generate_invoice_pdf
from datetime import datetime, date
import uuid
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import io
from decimal import Decimal
import decimal

# Helper functions for form field handling
def safe_decimal(value, default='0'):
    """Safely convert a value to Decimal with error handling"""
    try:
        if value is None or value == '':
            return Decimal(default)
        # Clean the value - remove any non-numeric characters except decimal point
        cleaned_value = str(value).strip()
        if not cleaned_value:
            return Decimal(default)
        return Decimal(cleaned_value)
    except (ValueError, TypeError, decimal.InvalidOperation):
        return Decimal(default)

def get_field_value(field):
    """Extract value from form field, handling both field objects and direct values"""
    if hasattr(field, 'data'):
        return field.data
    return field

def role_required(role):
    """Decorator for role-based access control"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login', next=request.url))
                
            if role == 'admin' and not hasattr(current_user, 'is_admin'):
                flash('Access denied. Administrator privileges required.', 'error')
                return redirect(url_for('dashboard'))
                
            if role == 'admin' and not current_user.is_admin():
                flash('Access denied. Administrator privileges required.', 'error')
                return redirect(url_for('dashboard'))
                
            if role == 'manager' and not hasattr(current_user, 'is_manager'):
                flash('Access denied. Manager privileges required.', 'error')
                return redirect(url_for('dashboard'))
                
            if role == 'manager' and not current_user.is_manager():
                flash('Access denied. Manager privileges required.', 'error')
                return redirect(url_for('dashboard'))
                
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# Authentication Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    # Redirect if already logged in
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data):
            # Check if user is active
            if not user.is_active:
                flash('Your account is deactivated. Please contact administrator.', 'danger')
                return render_template('login.html', title='Login', form=form)
            
            # Log the user in
            login_user(user, remember=form.remember_me.data if hasattr(form, 'remember_me') else False)
            
            # Redirect to next page if specified, otherwise index (dashboard)
            next_page = request.args.get('next')
            flash('Login successful!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        
        # Invalid credentials
        flash('Invalid username or password', 'danger')
    
    # Show login form
    return render_template('login.html', title='Login', form=form)

@app.route('/logout')
@login_required
def logout():
    """User logout"""
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def dashboard():
    """Dashboard showing recent activity and quick stats"""
    total_customers = Customer.query.count()
    total_products = Product.query.count()
    total_bills = Bill.query.count()
    recent_bills = Bill.query.order_by(Bill.created_at.desc()).limit(5).all()
    
    return render_template('index.html', 
                         total_customers=total_customers,
                         total_products=total_products,
                         total_bills=total_bills,
                         recent_bills=recent_bills)

@app.route('/company', methods=['GET', 'POST'])
@login_required
@role_required('manager')
def company_config():
    """Company configuration page"""
    company = Company.query.first()
    form = CompanyConfigForm(obj=company)
    
    if form.validate_on_submit():
        if not company:
            company = Company()
            db.session.add(company)
        
        form.populate_obj(company)
        
        # Handle logo upload
        if form.logo.data:
            filename = secure_filename(form.logo.data.filename)
            if filename and allowed_file(filename):
                # Generate unique filename
                file_ext = filename.rsplit('.', 1)[1].lower()
                unique_filename = f"{uuid.uuid4()}.{file_ext}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                
                # Create upload directory if it doesn't exist
                os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                
                # Delete old logo file if it exists
                if company.logo_path:
                    old_logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo_path)
                    if os.path.exists(old_logo_path):
                        try:
                            os.remove(old_logo_path)
                        except:
                            pass  # Ignore errors during removal
                
                form.logo.data.save(filepath)
                company.logo_path = unique_filename
        
        db.session.commit()
        flash('Company configuration updated successfully!', 'success')
        return redirect(url_for('company_config'))
    
    # Check if logo file exists
    if company and company.logo_path:
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], company.logo_path)
        if not os.path.exists(logo_path):
            flash('Warning: Logo file not found on disk.', 'warning')
    
    return render_template('company_config.html', form=form, company=company)

@app.route('/customers')
@login_required
def customers():
    """List all customers"""
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    
    query = Customer.query
    if search:
        query = query.filter(Customer.name.contains(search) | 
                           Customer.email.contains(search) |
                           Customer.gst_number.contains(search))
    
    customers = query.order_by(Customer.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    return render_template('customers.html', customers=customers, search=search)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    """Add new customer"""
    form = CustomerForm()
    
    if form.validate_on_submit():
        customer = Customer()
        form.populate_obj(customer)
        
        db.session.add(customer)
        db.session.commit()
        
        flash('Customer added successfully!', 'success')
        return redirect(url_for('customers'))
    
    return render_template('add_customer.html', form=form)

@app.route('/customers/<int:id>/edit', methods=['GET', 'POST'])
def edit_customer(id):
    """Edit existing customer"""
    customer = Customer.query.get_or_404(id)
    form = CustomerForm(obj=customer)
    
    if form.validate_on_submit():
        form.populate_obj(customer)
        db.session.commit()
        
        flash('Customer updated successfully!', 'success')
        return redirect(url_for('customers'))
    
    return render_template('add_customer.html', form=form, customer=customer)

@app.route('/customers/<int:id>/delete', methods=['POST'])
def delete_customer(id):
    """Delete customer"""
    customer = Customer.query.get_or_404(id)
    
    # Check if customer has bills
    if customer.bills:
        flash('Cannot delete customer with existing bills!', 'danger')
    else:
        db.session.delete(customer)
        db.session.commit()
        flash('Customer deleted successfully!', 'success')
    
    return redirect(url_for('customers'))

@app.route('/products')
@login_required
def products():
    """List all products"""
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    
    query = Product.query
    if search:
        query = query.filter(Product.name.contains(search) | 
                           Product.hsn_code.contains(search))
    
    products = query.order_by(Product.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    return render_template('products.html', products=products, search=search)

@app.route('/products/add', methods=['GET', 'POST'])
def add_product():
    """Add new product"""
    form = ProductForm()
    
    # Populate category choices
    categories = Category.query.order_by(Category.category_name).all()
    form.category.choices = [(0, 'None')] + [(c.id, c.category_name) for c in categories]
    
    # Get all enabled custom fields
    from models import FieldDefinition
    custom_fields = FieldDefinition.query.filter_by(
        entity_type='product', 
        enabled=True
    ).order_by(FieldDefinition.field_order).all()
    
    print("Request method:", request.method)  # Debug print
    if request.method == 'POST':
        print("Form data received:", request.form)  # Debug print
        
    # Add custom fields to the form
    form.add_custom_fields()
    
    if form.validate_on_submit():
        print("Form validated successfully")  # Debug print
        print("Form data after validation:", form.data)  # Debug print
        
        product = Product()
        
        # Copy basic fields
        for field in ['name', 'description', 'price', 'hsn_code', 'gst_rate', 'unit']:
            if hasattr(form, field):
                value = getattr(form, field).data
                print(f"Setting {field} = {value}")  # Debug print
                setattr(product, field, value)
        
        # Handle category separately
        if form.category.data and form.category.data != '0':
            category = Category.query.get(form.category.data)
            product.category = category
            print(f"Setting category: {category.category_name if category else 'None'}")  # Debug print
        else:
            product.category = None
            print("No category selected")  # Debug print
            
        # Initialize custom_fields
        product.custom_fields = {}
        
        # Save custom field values
        print("Raw form data:", request.form)  # Debug print
        
        new_custom_fields = {}  # Create a new dict for custom fields
        
        # Process all form fields that start with 'custom_'
        for key, value in request.form.items():
            if key.startswith('custom_'):
                # Remove the 'custom_' prefix to get the actual field name
                field_name = key[7:]  # Skip 'custom_'
                print(f"Processing {key}: value = {value}")  # Debug print
                
                # Find the field definition to get the correct type
                field_def = next((f for f in custom_fields if f.field_name == field_name), None)
                if field_def:
                    # Convert the value based on field type
                    if field_def.field_type == 'number':
                        try:
                            # Handle empty string
                            if value.strip() == '':
                                processed_value = None
                            else:
                                processed_value = float(value)
                        except ValueError:
                            processed_value = None
                    elif field_def.field_type == 'boolean':
                        processed_value = value.lower() in ('true', '1', 'yes', 'on')
                    else:
                        # For text fields, store empty string as None
                        processed_value = value if value.strip() != '' else None
                    
                    if processed_value is not None:
                        print(f"Saving {field_name} = {processed_value} ({type(processed_value)})")  # Debug print
                        new_custom_fields[field_name] = processed_value
        
        print("Final custom fields to save:", new_custom_fields)  # Debug print
        product.custom_fields = new_custom_fields
        
        db.session.add(product)
        db.session.commit()
        
        flash('Product added successfully!', 'success')
        return redirect(url_for('products'))
    
    return render_template('add_product.html', form=form, custom_fields=custom_fields)

@app.route('/products/<int:id>/edit', methods=['GET', 'POST'])
def edit_product(id):
    """Edit existing product"""
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    
    # Populate category choices
    categories = Category.query.order_by(Category.category_name).all()
    form.category.choices = [(0, 'None')] + [(c.id, c.category_name) for c in categories]

    # Set the selected category using the correct field
    form.category.data = str(product.category.id if product.category else 0)
    
    # Prepare custom field values before adding fields to form
    # Add custom fields to the form and pass the product's custom fields directly
    form.custom_field_values = product.custom_fields if product.custom_fields else {}
    print("Setting custom field values:", form.custom_field_values)  # Debug print
    
    # Add custom fields to the form
    custom_fields = form.add_custom_fields()

    # Print all form fields and their values for debugging
    print("All form fields:", {name: field.data for name, field in form._fields.items()})
    
    if form.validate_on_submit():
        # Copy basic fields
        for field in ['name', 'description', 'price', 'hsn_code', 'gst_rate', 'unit']:
            if hasattr(form, field):
                setattr(product, field, getattr(form, field).data)
        
        # Handle category separately
        category_id = int(form.category.data) if form.category.data else 0
        if category_id > 0:
            category = Category.query.get(category_id)
            product.category = category
        else:
            product.category = None
            
        # Initialize custom_fields JSON if it doesn't exist
        if not product.custom_fields:
            product.custom_fields = {}
        
        # Save custom field values
        print("Before save - existing custom_fields:", product.custom_fields)  # Debug print
        print("Raw form data:", request.form)  # Debug print of raw form data
        
        new_custom_fields = {}  # Create a new dict instead of modifying in place
        
        # Process all form fields that start with 'custom_'
        for key, value in request.form.items():
            if key.startswith('custom_'):
                # Remove the 'custom_' prefix to get the actual field name
                field_name = key[7:]  # Skip 'custom_'
                print(f"Processing {key}: value = {value}")  # Debug print
                
                # Find the field definition to get the correct type
                field_def = next((f for f in custom_fields if f.field_name == field_name), None)
                if field_def:
                    # Convert the value based on field type
                    if field_def.field_type == 'number':
                        try:
                            # Handle empty string
                            if value.strip() == '':
                                processed_value = None
                            else:
                                processed_value = float(value)
                        except ValueError:
                            processed_value = None
                    elif field_def.field_type == 'boolean':
                        processed_value = value.lower() in ('true', '1', 'yes', 'on')
                    else:
                        # For text fields, store empty string as None
                        processed_value = value if value.strip() != '' else None
                    
                    if processed_value is not None:
                        print(f"Saving {field_name} = {processed_value} ({type(processed_value)})")  # Debug print
                        new_custom_fields[field_name] = processed_value
                    
        print("Final custom fields to save:", new_custom_fields)  # Debug print
        product.custom_fields = new_custom_fields  # Assign the new dict all at once
        
        print("After save - new custom_fields:", product.custom_fields)  # Debug print
        db.session.commit()
        
        flash('Product updated successfully!', 'success')
        return redirect(url_for('products'))
    
    return render_template('add_product.html', form=form, product=product, custom_fields=custom_fields)

@app.route('/products/<int:id>/delete', methods=['POST'])
def delete_product(id):
    """Delete product"""
    product = Product.query.get_or_404(id)
    
    db.session.delete(product)
    db.session.commit()
    flash('Product deleted successfully!', 'success')
    
    return redirect(url_for('products'))

@app.route('/api/products/quick-add', methods=['POST'])
@login_required
def quick_add_product():
    """API endpoint for quickly adding a product during bill creation"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['name', 'price', 'hsn_code', 'gst_rate', 'unit']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'error': f'{field} is required'}), 400
        
        # Check if product already exists
        existing_product = Product.query.filter_by(name=data['name']).first()
        if existing_product:
            return jsonify({'success': False, 'error': 'Product with this name already exists'}), 400
        
        # Create new product
        product = Product(
            name=data['name'],
            description=data.get('description', ''),
            price=safe_decimal(data['price'], '0'),
            hsn_code=data['hsn_code'],
            gst_rate=safe_decimal(data['gst_rate'], '18'),
            unit=data['unit'],
            category=data.get('category', 'General')
        )
        
        db.session.add(product)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'description': product.description,
                'price': float(product.price),
                'hsn_code': product.hsn_code,
                'gst_rate': float(product.gst_rate),
                'unit': product.unit,
                'category': product.category
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/bills')
@login_required
def bills():
    """List all bills with filtering and export options"""
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    
    query = Bill.query.join(Customer)
    
    # Text search
    if search:
        query = query.filter(Bill.bill_number.contains(search) |
                           Customer.name.contains(search))
    
    # Date range filter
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date >= start_date_obj)
        except ValueError:
            flash('Invalid start date format', 'danger')
            start_date = ''
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date <= end_date_obj)
        except ValueError:
            flash('Invalid end date format', 'danger')
            end_date = ''
    
    # Status filter
    if status:
        query = query.filter(Bill.status == status)
    
    # Get summary statistics for the filtered results
    total_bills = query.count()
    total_amount = query.with_entities(db.func.sum(Bill.total_amount)).scalar() or 0
    
    bills = query.order_by(Bill.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    return render_template('bills.html', 
                         bills=bills, 
                         search=search,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         total_bills=total_bills,
                         total_amount=total_amount)

@app.route('/bills/export/excel')
@login_required
def export_bills_excel():
    """Export filtered bills to Excel"""
    search = request.args.get('search', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    
    # Build query with same filters as bills view
    query = Bill.query.join(Customer)
    
    if search:
        query = query.filter(Bill.bill_number.contains(search) |
                           Customer.name.contains(search))
    
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date <= end_date_obj)
        except ValueError:
            pass
    
    if status:
        query = query.filter(Bill.status == status)
    
    bills = query.order_by(Bill.created_at.desc()).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills Export"
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Header row
    headers = ['Bill Number', 'Customer Name', 'Bill Date', 'Due Date', 'Status', 
               'Subtotal', 'CGST', 'SGST', 'IGST', 'Total Amount']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, bill in enumerate(bills, 2):
        ws.cell(row=row, column=1, value=bill.bill_number)
        ws.cell(row=row, column=2, value=bill.customer.name)
        ws.cell(row=row, column=3, value=bill.bill_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=bill.due_date.strftime('%d/%m/%Y') if bill.due_date else '')
        ws.cell(row=row, column=5, value=bill.status)
        ws.cell(row=row, column=6, value=float(bill.subtotal))
        ws.cell(row=row, column=7, value=float(bill.cgst_amount))
        ws.cell(row=row, column=8, value=float(bill.sgst_amount))
        ws.cell(row=row, column=9, value=float(bill.igst_amount))
        ws.cell(row=row, column=10, value=float(bill.total_amount))
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    date_range = ""
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_range = f"_from_{start_date}"
    elif end_date:
        date_range = f"_until_{end_date}"
    
    filename = f"bills_export{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/bills/export/pdf')
@login_required
def export_bills_pdf():
    """Export filtered bills to PDF"""
    search = request.args.get('search', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    
    # Build query with same filters as bills view
    query = Bill.query.join(Customer)
    
    if search:
        query = query.filter(Bill.bill_number.contains(search) |
                           Customer.name.contains(search))
    
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date >= start_date_obj)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            query = query.filter(Bill.bill_date <= end_date_obj)
        except ValueError:
            pass
    
    if status:
        query = query.filter(Bill.status == status)
    
    bills = query.order_by(Bill.created_at.desc()).all()
    company = Company.query.first()
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    # Title
    title = f"Bills Export Report"
    if start_date and end_date:
        title += f" ({start_date} to {end_date})"
    elif start_date:
        title += f" (from {start_date})"
    elif end_date:
        title += f" (until {end_date})"
    
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    
    # Company info
    if company:
        elements.append(Paragraph(f"<b>{company.name}</b>", styles['Normal']))
        elements.append(Paragraph(f"GST No: {company.gst_number}", styles['Normal']))
        elements.append(Spacer(1, 12))
    
    # Summary
    total_amount = sum(float(bill.total_amount) for bill in bills)
    elements.append(Paragraph(f"<b>Total Bills:</b> {len(bills)}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Amount:</b> ₹{total_amount:,.2f}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table data
    data = [['Bill Number', 'Customer', 'Date', 'Status', 'Amount']]
    
    for bill in bills:
        data.append([
            bill.bill_number,
            bill.customer.name,
            bill.bill_date.strftime('%d/%m/%Y'),
            bill.status,
            f"₹{float(bill.total_amount):,.2f}"
        ])
    
    # Create table
    table = Table(data, colWidths=[100, 150, 80, 80, 100])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Right align amount column
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Generate filename
    date_range = ""
    if start_date and end_date:
        date_range = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_range = f"_from_{start_date}"
    elif end_date:
        date_range = f"_until_{end_date}"
    
    filename = f"bills_export{date_range}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@app.route('/bills/create', methods=['GET', 'POST'])
@login_required
def create_bill():
    """Create new bill"""
    form = BillForm()
    
    # Populate customer choices
    customers = Customer.query.order_by(Customer.name).all()
    form.customer_id.choices = [(c.id, f"{c.name} {'(Guest)' if c.is_guest else ''}") for c in customers]
    
    # Get all products for JavaScript and form choices
    products = Product.query.all()
    product_choices = [(0, 'Select Product')] + [(p.id, f"{p.name} - ₹{p.price}") for p in products]
    
    # Handle duplicate functionality
    duplicate_id = request.args.get('duplicate')
    duplicate_bill = None
    if duplicate_id:
        duplicate_bill = Bill.query.get(duplicate_id)
        if duplicate_bill:
            # Pre-fill form with duplicate data
            form.customer_id.data = duplicate_bill.customer_id
            form.notes.data = duplicate_bill.notes
            form.discount_type.data = duplicate_bill.discount_type or 'none'
            form.discount_value.data = duplicate_bill.discount_value or 0
    
    # Populate product choices for each item in the form
    for item_form in form.items:
        item_form.product_id.choices = product_choices
    
    # For GET requests, ensure we have at least one item with proper choices
    if request.method == 'GET':
        if not form.items.entries:
            form.items.append_entry()
        for item_form in form.items:
            item_form.product_id.choices = product_choices
    
    if form.validate_on_submit():
        # Generate bill number
        current_year = datetime.now().year
        sequence = BillSequence.query.filter_by(year=current_year).first()
        if not sequence:
            sequence = BillSequence(year=current_year, sequence_number=0)
            db.session.add(sequence)
        
        sequence.sequence_number += 1
        bill_number = f"INV-{current_year}-{sequence.sequence_number:04d}"
        
        # Create bill
        bill = Bill(
            bill_number=bill_number,
            customer_id=form.customer_id.data,
            bill_date=form.bill_date.data,
            due_date=form.due_date.data,
            discount_type=form.discount_type.data or 'none',
            discount_value=form.discount_value.data or 0,
            notes=form.notes.data
        )
        
        # Get customer and company for GST calculations
        customer = Customer.query.get(form.customer_id.data)
        company = Company.query.first()
        
        # Calculate totals
        from decimal import Decimal
        subtotal = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        
        # Process items
        for item_form in form.items:
            if item_form.product_name.data:
                # Handle the case where form fields might be strings or field objects
                def get_field_value(field):
                    if hasattr(field, 'data'):
                        return field.data
                    return field
                
                # Get values and ensure they're proper Decimal types
                quantity_val = get_field_value(item_form.quantity)
                rate_val = get_field_value(item_form.rate)
                gst_rate_val = get_field_value(item_form.gst_rate)
                
                item = BillItem(
                    product_name=get_field_value(item_form.product_name),
                    description=get_field_value(item_form.description) or '',
                    hsn_code=get_field_value(item_form.hsn_code),
                    quantity=safe_decimal(quantity_val, '1'),
                    unit=get_field_value(item_form.unit),
                    rate=safe_decimal(rate_val, '0'),
                    unit_price=safe_decimal(rate_val, '0'),  # Set unit_price same as rate
                    gst_rate=safe_decimal(gst_rate_val, '18')
                )
                
                # Calculate amounts
                item.amount = item.quantity * item.rate
                item.taxable_amount = item.amount  # Set taxable_amount same as amount for legacy support
                subtotal += item.amount
                
                # Calculate GST
                gst_amounts = calculate_gst(
                    item.amount, 
                    item.gst_rate, 
                    company.state_code if company else '27',
                    customer.state_code if customer and customer.state_code else company.state_code if company else '27'
                )
                
                item.cgst_amount = gst_amounts['cgst']
                item.sgst_amount = gst_amounts['sgst']
                item.igst_amount = gst_amounts['igst']

                # Set legacy fields
                item.gst_amount = item.cgst_amount + item.sgst_amount + item.igst_amount
                item.total_amount = item.amount + item.gst_amount

                
                total_cgst += item.cgst_amount
                total_sgst += item.sgst_amount
                total_igst += item.igst_amount
                
                bill.items.append(item)
        
        # Calculate discount
        discount_amount = Decimal('0')
        if bill.discount_type == 'percentage' and bill.discount_value > 0:
            discount_amount = (subtotal * bill.discount_value) / 100
        elif bill.discount_type == 'amount' and bill.discount_value > 0:
            discount_amount = min(bill.discount_value, subtotal)
        
        # Set bill totals
        bill.subtotal = subtotal
        bill.discount_amount = discount_amount
        bill.cgst_amount = total_cgst
        bill.sgst_amount = total_sgst
        bill.igst_amount = total_igst
        bill.gst_amount = total_cgst + total_sgst + total_igst  # Calculate total GST amount
        bill.total_amount = subtotal - discount_amount + total_cgst + total_sgst + total_igst
        bill.final_amount = bill.total_amount  # Set final_amount same as total_amount for legacy support
        
        db.session.add(bill)
        db.session.commit()
        
        flash('Bill created successfully!', 'success')
        return redirect(url_for('view_bill', id=bill.id))
    
    # Just get the most recent/popular products for initial view
    # Rest will be loaded via AJAX as needed
    recent_products = Product.query.order_by(Product.created_at.desc()).limit(20).all()
    
    # Convert products to JSON-serializable format
    products_data = []
    for p in recent_products:
        products_data.append({
            'id': p.id,
            'name': p.name,
            'description': p.description,
            'price': float(p.price),
            'hsn_code': p.hsn_code,
            'gst_rate': float(p.gst_rate),
            'unit': p.unit,
            'category': p.category.category_name if p.category else 'General'
        })
    
    return render_template('create_bill.html', form=form, products=products_data, duplicate_bill=duplicate_bill)

@app.route('/bills/<int:id>', methods=['GET', 'POST'])
@login_required
def view_bill(id):
    """View bill details and handle status updates"""
    bill = Bill.query.get_or_404(id)
    company = Company.query.first()
    
    # Handle POST request for status update
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'update_status':
            new_status = request.form.get('status')
            if new_status in ['Draft', 'Sent', 'Paid', 'Cancelled']:
                bill.status = new_status
                bill.updated_at = datetime.now()
                db.session.commit()
                flash(f'Bill status updated to {new_status}!', 'success')
            else:
                flash('Invalid status selected!', 'error')
        return redirect(url_for('view_bill', id=id))
    
    current_time = datetime.now().strftime('%d/%m/%Y %I:%M %p')
    
    return render_template('view_bill.html', bill=bill, company=company, current_time=current_time)

@app.route('/bills/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bill(id):
    """Edit existing bill"""
    bill = Bill.query.get_or_404(id)
    form = BillForm(obj=bill)
    
    # Populate customer choices
    customers = Customer.query.order_by(Customer.name).all()
    form.customer_id.choices = [(c.id, f"{c.name} {'(Guest)' if c.is_guest else ''}") for c in customers]
    
    # Get all products for JavaScript and form choices
    products = Product.query.all()
    product_choices = [(0, 'Select Product')] + [(p.id, f"{p.name} - ₹{p.price}") for p in products]
    
    # Clear existing items and populate with current bill items
    form.items.entries.clear()
    for item in bill.items:
        item_form = BillItemForm()
        item_form.product_id.choices = product_choices
        item_form.product_id.data = item.product_id if item.product_id else 0
        item_form.product_name.data = item.product_name
        item_form.description.data = item.description
        item_form.hsn_code.data = item.hsn_code
        item_form.quantity.data = item.quantity
        item_form.unit.data = item.unit
        item_form.rate.data = item.rate
        item_form.gst_rate.data = item.gst_rate
        form.items.append_entry(item_form)
    
    # Populate product choices for each item in the form
    for item_form in form.items:
        item_form.product_id.choices = product_choices
    
    if form.validate_on_submit():
        # Update bill details
        bill.customer_id = form.customer_id.data
        bill.bill_date = form.bill_date.data
        bill.due_date = form.due_date.data
        bill.status = form.status.data
        bill.discount_type = form.discount_type.data or 'none'
        bill.discount_value = form.discount_value.data or 0
        bill.notes = form.notes.data
        
        # Clear existing items
        for item in bill.items:
            db.session.delete(item)
        
        # Get customer and company for GST calculations
        customer = Customer.query.get(form.customer_id.data)
        company = Company.query.first()
        
        # Calculate totals
        from decimal import Decimal
        subtotal = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        
        # Process items
        for item_form in form.items:
            if item_form.product_name.data:
                # Handle the case where form fields might be strings or field objects
                def get_field_value(field):
                    if hasattr(field, 'data'):
                        return field.data
                    return field
                
                # Get values and ensure they're proper Decimal types
                quantity_val = get_field_value(item_form.quantity)
                rate_val = get_field_value(item_form.rate)
                gst_rate_val = get_field_value(item_form.gst_rate)
                
                item = BillItem(
                    product_name=get_field_value(item_form.product_name),
                    description=get_field_value(item_form.description) or '',
                    hsn_code=get_field_value(item_form.hsn_code),
                    quantity=safe_decimal(quantity_val, '1'),
                    unit=get_field_value(item_form.unit),
                    rate=safe_decimal(rate_val, '0'),
                    unit_price=safe_decimal(rate_val, '0'),  # Set unit_price same as rate
                    gst_rate=safe_decimal(gst_rate_val, '18')
                )
                
                # Calculate amounts
                item.amount = item.quantity * item.rate
                item.taxable_amount = item.amount  # Set taxable_amount same as amount for legacy support
                subtotal += item.amount
                
                # Calculate GST
                gst_amounts = calculate_gst(
                    item.amount, 
                    item.gst_rate, 
                    company.state_code if company else '27',
                    customer.state_code if customer and customer.state_code else company.state_code if company else '27'
                )
                
                item.cgst_amount = gst_amounts['cgst']
                item.sgst_amount = gst_amounts['sgst']
                item.igst_amount = gst_amounts['igst']

                # Set legacy fields
                item.gst_amount = item.cgst_amount + item.sgst_amount + item.igst_amount
                item.total_amount = item.amount + item.gst_amount

                
                total_cgst += item.cgst_amount
                total_sgst += item.sgst_amount
                total_igst += item.igst_amount
                
                bill.items.append(item)
        
        # Calculate discount for edit bill
        discount_amount = Decimal('0')
        if bill.discount_type == 'percentage' and bill.discount_value > 0:
            discount_amount = (subtotal * bill.discount_value) / 100
        elif bill.discount_type == 'amount' and bill.discount_value > 0:
            discount_amount = min(bill.discount_value, subtotal)
        
        # Set bill totals
        bill.subtotal = subtotal
        bill.discount_amount = discount_amount
        bill.cgst_amount = total_cgst
        bill.sgst_amount = total_sgst
        bill.igst_amount = total_igst
        bill.gst_amount = total_cgst + total_sgst + total_igst  # Calculate total GST amount
        bill.total_amount = subtotal - discount_amount + total_cgst + total_sgst + total_igst
        bill.final_amount = bill.total_amount  # Set final_amount same as total_amount for legacy support
        
        # Update modified timestamp
        bill.updated_at = datetime.now()
        
        db.session.commit()
        
        flash('Bill updated successfully!', 'success')
        return redirect(url_for('view_bill', id=bill.id))
    
    # Get essential product data for initialization
    essential_products = []
    
    # Include products that are in the current bill
    bill_product_ids = [item.product_id for item in bill.items if item.product_id]
    if bill_product_ids:
        bill_products = Product.query.filter(Product.id.in_(bill_product_ids)).all()
        essential_products.extend(bill_products)
    
    # Add a few recent products to provide initial options
    recent_count = 20 - len(essential_products)
    if recent_count > 0:
        recent_products = Product.query.filter(~Product.id.in_(bill_product_ids)) \
            .order_by(Product.created_at.desc()) \
            .limit(recent_count).all()
        essential_products.extend(recent_products)
    
    # Convert products to JSON-serializable format
    products_data = []
    for p in essential_products:
        products_data.append({
            'id': p.id,
            'name': p.name,
            'description': p.description,
            'price': float(p.price),
            'hsn_code': p.hsn_code,
            'gst_rate': float(p.gst_rate),
            'unit': p.unit,
            'category': p.category.category_name if p.category else 'General'
        })
    
    return render_template('create_bill.html', form=form, products=products_data, edit_mode=True, bill=bill)

@app.route('/bills/<int:id>/pdf')
def download_bill_pdf(id):
    """Download bill as PDF"""
    bill = Bill.query.get_or_404(id)
    company = Company.query.first()
    
    if not company:
        flash('Please configure company details first!', 'warning')
        return redirect(url_for('company_config'))
    
    pdf_file = generate_invoice_pdf(bill, company)
    
    return send_file(
        pdf_file,
        as_attachment=True,
        download_name=f"Invoice_{bill.bill_number}.pdf",
        mimetype='application/pdf'
    )

@app.route('/api/products/<int:id>')
def get_product_api(id):
    """API endpoint to get product details"""
    product = Product.query.get_or_404(id)
    return jsonify({
        'product': {
            'id': product.id,
            'name': product.name,
            'description': product.description,
            'price': float(product.price),
            'hsn_code': product.hsn_code,
            'gst_rate': float(product.gst_rate),
            'unit': product.unit,
            'category': product.category.name if product.category else 'General'
        }
    })
    
@app.route('/api/products/search')
@login_required
def search_products_api():
    """API endpoint to search for products"""
    term = request.args.get('term', '').strip()
    category = request.args.get('category', '')
    
    # Base query
    query = Product.query
    
    # Apply search term filter
    if term:
        query = query.filter(
            db.or_(
                Product.name.ilike(f"%{term}%"),
                Product.description.ilike(f"%{term}%"),
                Product.hsn_code.ilike(f"%{term}%")
            )
        )
    
    # Apply category filter
    if category:
        query = query.join(Category).filter(Category.name == category)
    
    # Get results (limit to 100 for performance)
    products = query.order_by(Product.name).limit(100).all()
    
    # Format results
    result = {
        'products': [
            {
                'id': p.id,
                'name': p.name,
                'description': p.description,
                'price': float(p.price),
                'hsn_code': p.hsn_code,
                'gst_rate': float(p.gst_rate),
                'unit': p.unit,
                'category': p.category.category_name if p.category else 'General'
            } for p in products
        ]
    }
    
    return jsonify(result)

@app.route('/api/products/recent')
@login_required
def recent_products_api():
    """API endpoint to get recent/popular products"""
    # Get recently created products
    recent_products = Product.query.order_by(Product.created_at.desc()).limit(20).all()
    
    # Format results
    result = {
        'products': [
            {
                'id': p.id,
                'name': p.name,
                'description': p.description,
                'price': float(p.price),
                'hsn_code': p.hsn_code,
                'gst_rate': float(p.gst_rate),
                'unit': p.unit,
                'category': p.category.category_name if p.category else 'General'
            } for p in recent_products
        ]
    }
    
    return jsonify(result)

@app.route('/api/customers/quick-add', methods=['POST'])
def quick_add_customer():
    """API endpoint to quickly add a guest customer"""
    data = request.get_json()
    
    customer = Customer(
        name=data.get('name'),
        email=data.get('email'),
        phone=data.get('phone'),
        address=data.get('address'),
        is_guest=True
    )
    
    db.session.add(customer)
    db.session.commit()
    
    return jsonify({
        'id': customer.id,
        'name': customer.name,
        'success': True
    })

# User Management Routes
@app.route('/users')
@login_required
@role_required('admin')
def users():
    """List all users"""
    search = request.args.get('search', '')
    page = request.args.get('page', 1, type=int)
    
    query = User.query
    if search:
        query = query.filter(User.username.contains(search) | 
                           User.email.contains(search) |
                           User.first_name.contains(search) |
                           User.last_name.contains(search))
    
    users = query.order_by(User.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False)
    
    return render_template('users.html', users=users, search=search)

@app.route('/users/add', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def add_user():
    """Add new user"""
    form = CreateUserForm()
    
    if form.validate_on_submit():
        # Check if username or email already exists
        existing_user = User.query.filter(
            (User.username == form.username.data) | 
            (User.email == form.email.data)
        ).first()
        
        if existing_user:
            if existing_user.username == form.username.data:
                flash('Username already exists!', 'error')
            else:
                flash('Email already exists!', 'error')
        else:
            user = User()
            form.populate_obj(user)
            user.set_password(form.password.data)
            
            db.session.add(user)
            db.session.commit()
            
            flash(f'User {user.username} created successfully!', 'success')
            return redirect(url_for('users'))
    
    return render_template('add_user.html', form=form)

@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def edit_user(id):
    """Edit existing user"""
    user = User.query.get_or_404(id)
    form = UserForm(obj=user)
    
    if form.validate_on_submit():
        # Check if username or email already exists (excluding current user)
        existing_user = User.query.filter(
            ((User.username == form.username.data) | 
             (User.email == form.email.data)) &
            (User.id != user.id)
        ).first()
        
        if existing_user:
            if existing_user.username == form.username.data:
                flash('Username already exists!', 'error')
            else:
                flash('Email already exists!', 'error')
        else:
            form.populate_obj(user)
            
            # Update password if provided
            if form.password.data:
                user.set_password(form.password.data)
            
            db.session.commit()
            flash(f'User {user.username} updated successfully!', 'success')
            return redirect(url_for('users'))
    
    return render_template('add_user.html', form=form, user=user)

@app.route('/users/<int:id>/delete')
@login_required
@role_required('admin')
def delete_user(id):
    """Delete user"""
    user = User.query.get_or_404(id)
    
    # Prevent deleting own account
    if user.id == current_user.id:
        flash('You cannot delete your own account!', 'error')
        return redirect(url_for('users'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    flash(f'User {username} deleted successfully!', 'success')
    return redirect(url_for('users'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    """User profile page"""
    form = ChangePasswordForm()
    
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            db.session.commit()
            flash('Password changed successfully!', 'success')
            return redirect(url_for('profile'))
        else:
            flash('Current password is incorrect!', 'error')
    
    return render_template('profile.html', form=form)

# Category Routes
@app.route('/categories', methods=['GET', 'POST'])
@login_required
@role_required('manager')
def manage_categories():
    """Manage categories: list and add new categories"""
    from models import Category
    from forms import CategoryForm

    form = CategoryForm()
    categories = Category.query.order_by(Category.created_at.desc()).all()

    if form.validate_on_submit():
        new_category = Category(category_name=form.category_name.data)
        db.session.add(new_category)
        db.session.commit()
        flash('Category added successfully!', 'success')
        return redirect(url_for('manage_categories'))

    return render_template('categories.html', form=form, categories=categories)

@app.route('/categories/edit/<int:category_id>', methods=['GET', 'POST'])
@login_required
@role_required('manager')
def edit_category(category_id):
    """Edit an existing category"""
    from models import Category
    from forms import CategoryForm

    category = Category.query.get_or_404(category_id)
    form = CategoryForm(obj=category)

    if form.validate_on_submit():
        category.category_name = form.category_name.data
        db.session.commit()
        flash('Category updated successfully!', 'success')
        return redirect(url_for('manage_categories'))

    return render_template('edit_category.html', form=form, category=category)

@app.route('/categories/delete/<int:category_id>', methods=['POST'])
@login_required
@role_required('manager')
def delete_category(category_id):
    """Delete an existing category"""
    from models import Category

    category = Category.query.get_or_404(category_id)
    db.session.delete(category)
    db.session.commit()
    flash('Category deleted successfully!', 'success')
    return redirect(url_for('manage_categories'))

# Session and CSRF handling
@app.before_request
def make_session_permanent():
    session.permanent = True

@app.route('/csrf-token')
def csrf_token():
    """Get CSRF token for AJAX requests"""
    return jsonify({'csrf_token': generate_csrf()})

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve uploaded files"""
    return send_file(os.path.join(app.config['UPLOAD_FOLDER'], filename))

@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500
